'''
Steps :
1) Get a CA Number from workbook.
2) Get all the related fields required for submission of a complaint from the CA Number.
3) Using these fields and CA Number lodge a complaint. [Also get a random message from a list of messages to submit]
4) Save the complaint number with timestamp to the sheet.
5) Repeat 1 to 4 for all CA Numbers [First check for one CA number only]
6) Show a sucess message
7) Exit gracefully
'''

'''
Resources :
-Link for complaint submission - https://www.tatapower-ddl.com/customer/complaint/regcomp_detailsrevamp.aspx
-Accepts a POST request ONLY!
-There can be two types POST requests

1) To get details of CA number

    getCADetailsPayload = {
    '__VIEWSTATE' : '/wEPDwUILTg3MjIxMDYPZBYCAgEPZBYaZg8PFgIeBFRleHQFgwFPcHRpb24gU2VsZWN0ZWQgOiA8dT5ObyBQb3dlcjwvdT4gKE8mTSAgQ29tcGxhaW50KS4gPGJyPjxhIGhyZWY9J2NvbXBsYWludC1yZWdpc3RyYXRpb24uYXNweCc+Q2xpY2sgaGVyZTwvYT4gdG8gc2VsZWN0IG90aGVyIG9wdGlvbmRkAgIPDxYCHgdWaXNpYmxlZ2RkAgMPDxYEHglCYWNrQ29sb3IKIR4EXyFTQgIIZGQCBQ8PFgIfAWdkZAIGDw8WBh8CCiEeCFJlYWRPbmx5Zx8DAghkZAIHDw9kFgIeCm9ua2V5cHJlc3MFJmphdmFzY3JpcHQ6cmV0dXJuIGFsbG93bnVtYmVycyhldmVudCk7ZAIIDw9kFgIfBQUmamF2YXNjcmlwdDpyZXR1cm4gYWxsb3dudW1iZXJzKGV2ZW50KTtkAgkPDxYCHwFnZGQCCg8PFgYfAgohHwRnHwMCCGRkAg0PEGRkFgBkAg8PDxYCHwFnZGQCEA8PFgQfAgohHwMCCGRkAhEPDxYEHgdUb29sVGlwBRdJbWFnZSBWZXJpZmljYXRpb24gQ29kZR4ISW1hZ2VVcmwFCn4vSW1nLmFzcHhkZGR8HzDRmY0tnIDT84vqDDTUVfzuhoOXvx04gBQKfM+N5Q==',
    '__VIEWSTATEGENERATOR' : 'FA96F2E7',
    '__EVENTVALIDATION' : '/wEdAApYaIhDCcMW/ET5/xEE7sC2EcOgvHRMMdknX1hKEqsA4Zysd+iXqvJbUqR3wDY/ZCBqaqe4WPrPNXSx0nLeM0MxIkjkwO2Hn9+IKdzB92ZTmTEmrsU+Uo4cUGFL28mNcdDPnskGUk9pqqILy5LAvTehaihfy87szX5n2LyriLG2osZV+oV9ZHJPsVyMonIqKvYfeetr3c4H3XmikgJc2ZlI/ecVh9iV+ytFuVNnQebcaFAdHDEasYEiOT9ymVIZlK8=',
    'searchtext' :'',
    'txtcano' : CANumber, #Only this field needs to change
    'btngetconsdetails' : 'Get Details',
    'txtconsname' :'',
    'txtphone' :'',
    'txtmobile' :'',
    'txtsuppadd' :'',
    'txtremarks' :'',
    'TxtImgVer' : 'Enter code as it appears'
    }

2) To lodge a complaint

Data sent -
* : mandatory
^ : autofilled

1) CA Number*^
2) Name*^
3) Landline
4) Mobile*
5) Address*^
6) Remarks*
7) Captcha code*

complaintRegisterPayload = {
'__VIEWSTATE':'/wEPDwUILTg3MjIxMDYPZBYCAgEPZBYcZg8PFgIeBFRleHQFgwFPcHRpb24gU2VsZWN0ZWQgOiA8dT5ObyBQb3dlcjwvdT4gKE8mTSAgQ29tcGxhaW50KS4gPGJyPjxhIGhyZWY9J2NvbXBsYWludC1yZWdpc3RyYXRpb24uYXNweCc+Q2xpY2sgaGVyZTwvYT4gdG8gc2VsZWN0IG90aGVyIG9wdGlvbmRkAgEPDxYCHwBlZGQCAg8PFgIeB1Zpc2libGVnZGQCAw8PFgQeCUJhY2tDb2xvcgohHgRfIVNCAghkZAIFDw8WAh8BZ2RkAgYPDxYIHwIKIR4IUmVhZE9ubHlnHwMCCB8ABRhNci4gUkFNRVNIV0FSIERBU1MgQk9IRVRkZAIHDw9kFgIeCm9ua2V5cHJlc3MFJmphdmFzY3JpcHQ6cmV0dXJuIGFsbG93bnVtYmVycyhldmVudCk7ZAIIDw9kFgIfBQUmamF2YXNjcmlwdDpyZXR1cm4gYWxsb3dudW1iZXJzKGV2ZW50KTtkAgkPDxYCHwFnZGQCCg8PFggfAgohHwRnHwMCCB8ABYkBUExPVCBOTyAxOSBLSC4gTk8uIDM4LzIyIEdST1VORCBGTE9PUiBHUkFNIFNBQkhBIFNUUkVFVCBOTy4gR0FMSSBOTy0yIFZJTExBR0UgTUFNVVJQVVIgQ0lUWSBERUxISSAxMTAwNDAgTEFORE1BUksgTkVBUiBSQURIQSBTV0FNSSBBU0hSQU1kZAINDxBkZBYAZAIPDw8WAh8BZ2RkAhAPDxYEHwIKIR8DAghkZAIRDw8WBB4HVG9vbFRpcAUXSW1hZ2UgVmVyaWZpY2F0aW9uIENvZGUeCEltYWdlVXJsBQp+L0ltZy5hc3B4ZGRkjxa9v723bAVIOcwU32DuAnlJ/CI7/HBtF1dnNqSutcQ=',
'__VIEWSTATEGENERATOR':'FA96F2E7',
'__EVENTVALIDATION':'/wEdAAo+Yg5UIlmHX9pXTkoEn0B9EcOgvHRMMdknX1hKEqsA4Zysd+iXqvJbUqR3wDY/ZCBqaqe4WPrPNXSx0nLeM0MxIkjkwO2Hn9+IKdzB92ZTmTEmrsU+Uo4cUGFL28mNcdDPnskGUk9pqqILy5LAvTehaihfy87szX5n2LyriLG2osZV+oV9ZHJPsVyMonIqKvYfeetr3c4H3XmikgJc2ZlI+O/nW7ap4phzpExocMkEqc8lYKgM3TV2apVtDd0IKaE=',
'searchtext':'',
'txtcano':'60014463164',
'txtconsname':'Mr. RAMESHWAR DASS BOHET',
'txtphone':'28520855',
'txtmobile':'9654476321',
'txtsuppadd':'PLOT NO 19 KH. NO. 38/22 GROUND FLOOR GRAM SABHA STREET NO. GALI NO-2 VILLAGE MAMURPUR CITY DELHI 110040 LANDMARK NEAR RADHA SWAMI ASHRAM',
'txtremarks': 'Remarks/Complaint details go Here',
'TxtImgVer':'Captcha field',
'btnsubmit': 'Submit'
}

'''
import openpyxl, requests, logging, random
from bs4 import BeautifulSoup
from datetime import datetime

logging.basicConfig(filename = 'logs.txt',level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
#logging.disable(logging.CRITICAL) #disables all logging messages

#File Name
file = 'CANumbers.xlsx' #This needs to be in PWD

#WorkSheet names
CANumbersWorkSheet = 'CANumbers'
complaintNumberWorkSheet = 'ComplaintNumbers'
remarksWorkSheet = 'Remarks'

#Opening workbook
workbook = openpyxl.load_workbook(filename = file) #Opening the .xlsx file
#TODO return an error if file is not found OR unable to open file.
logging.debug('Workbook opened')

#Worksheet objects
CASheet = workbook[CANumbersWorkSheet]
remarksSheet = workbook[remarksWorkSheet]
complaintSheet = workbook[complaintNumberWorkSheet]






def main():
    logging.debug('Started main()')

    NumberOfRecords = len(CASheet['A']) - 1
    
    for record in range(2, NumberOfRecords + 1):    #To skip the first line
        CANumber = CASheet.cell(row = record, column = 1).value
        
        if CANumber != None:
            payload = getDetailsFromCA(CANumber,record) #As of now sending index to update values in sheet
            lodgeComplaint(payload)
    print('Done!')
    logging.debug('Finished main()')
    
    
def getDetailsFromCA(CANumber,index):
    logging.debug('Started getDetailsFromCA()')
    
    getDetailsPayload = {
    '__VIEWSTATE' : '/wEPDwUILTg3MjIxMDYPZBYCAgEPZBYaZg8PFgIeBFRleHQFgwFPcHRpb24gU2VsZWN0ZWQgOiA8dT5ObyBQb3dlcjwvdT4gKE8mTSAgQ29tcGxhaW50KS4gPGJyPjxhIGhyZWY9J2NvbXBsYWludC1yZWdpc3RyYXRpb24uYXNweCc+Q2xpY2sgaGVyZTwvYT4gdG8gc2VsZWN0IG90aGVyIG9wdGlvbmRkAgIPDxYCHgdWaXNpYmxlZ2RkAgMPDxYEHglCYWNrQ29sb3IKIR4EXyFTQgIIZGQCBQ8PFgIfAWdkZAIGDw8WBh8CCiEeCFJlYWRPbmx5Zx8DAghkZAIHDw9kFgIeCm9ua2V5cHJlc3MFJmphdmFzY3JpcHQ6cmV0dXJuIGFsbG93bnVtYmVycyhldmVudCk7ZAIIDw9kFgIfBQUmamF2YXNjcmlwdDpyZXR1cm4gYWxsb3dudW1iZXJzKGV2ZW50KTtkAgkPDxYCHwFnZGQCCg8PFgYfAgohHwRnHwMCCGRkAg0PEGRkFgBkAg8PDxYCHwFnZGQCEA8PFgQfAgohHwMCCGRkAhEPDxYEHgdUb29sVGlwBRdJbWFnZSBWZXJpZmljYXRpb24gQ29kZR4ISW1hZ2VVcmwFCn4vSW1nLmFzcHhkZGR8HzDRmY0tnIDT84vqDDTUVfzuhoOXvx04gBQKfM+N5Q==',
    '__VIEWSTATEGENERATOR' : 'FA96F2E7',
    '__EVENTVALIDATION' : '/wEdAApYaIhDCcMW/ET5/xEE7sC2EcOgvHRMMdknX1hKEqsA4Zysd+iXqvJbUqR3wDY/ZCBqaqe4WPrPNXSx0nLeM0MxIkjkwO2Hn9+IKdzB92ZTmTEmrsU+Uo4cUGFL28mNcdDPnskGUk9pqqILy5LAvTehaihfy87szX5n2LyriLG2osZV+oV9ZHJPsVyMonIqKvYfeetr3c4H3XmikgJc2ZlI/ecVh9iV+ytFuVNnQebcaFAdHDEasYEiOT9ymVIZlK8=',
    'searchtext' :'',
    'txtcano' : CANumber,
    'btngetconsdetails' : 'Get Details',
    'txtconsname' :'',
    'txtphone' :'',
    'txtmobile' :'',
    'txtsuppadd' :'',
    'txtremarks' :'',
    'TxtImgVer' : 'Enter code as it appears'
    }

    complaintRegisterPayload = {
    '__VIEWSTATE':'/wEPDwUILTg3MjIxMDYPZBYCAgEPZBYcZg8PFgIeBFRleHQFgwFPcHRpb24gU2VsZWN0ZWQgOiA8dT5ObyBQb3dlcjwvdT4gKE8mTSAgQ29tcGxhaW50KS4gPGJyPjxhIGhyZWY9J2NvbXBsYWludC1yZWdpc3RyYXRpb24uYXNweCc+Q2xpY2sgaGVyZTwvYT4gdG8gc2VsZWN0IG90aGVyIG9wdGlvbmRkAgEPDxYCHwBlZGQCAg8PFgIeB1Zpc2libGVnZGQCAw8PFgQeCUJhY2tDb2xvcgohHgRfIVNCAghkZAIFDw8WAh8BZ2RkAgYPDxYIHwIKIR4IUmVhZE9ubHlnHwMCCB8ABRhNci4gUkFNRVNIV0FSIERBU1MgQk9IRVRkZAIHDw9kFgIeCm9ua2V5cHJlc3MFJmphdmFzY3JpcHQ6cmV0dXJuIGFsbG93bnVtYmVycyhldmVudCk7ZAIIDw9kFgIfBQUmamF2YXNjcmlwdDpyZXR1cm4gYWxsb3dudW1iZXJzKGV2ZW50KTtkAgkPDxYCHwFnZGQCCg8PFggfAgohHwRnHwMCCB8ABYkBUExPVCBOTyAxOSBLSC4gTk8uIDM4LzIyIEdST1VORCBGTE9PUiBHUkFNIFNBQkhBIFNUUkVFVCBOTy4gR0FMSSBOTy0yIFZJTExBR0UgTUFNVVJQVVIgQ0lUWSBERUxISSAxMTAwNDAgTEFORE1BUksgTkVBUiBSQURIQSBTV0FNSSBBU0hSQU1kZAINDxBkZBYAZAIPDw8WAh8BZ2RkAhAPDxYEHwIKIR8DAghkZAIRDw8WBB4HVG9vbFRpcAUXSW1hZ2UgVmVyaWZpY2F0aW9uIENvZGUeCEltYWdlVXJsBQp+L0ltZy5hc3B4ZGRkjxa9v723bAVIOcwU32DuAnlJ/CI7/HBtF1dnNqSutcQ=',
    '__VIEWSTATEGENERATOR':'FA96F2E7',
    '__EVENTVALIDATION':'/wEdAAo+Yg5UIlmHX9pXTkoEn0B9EcOgvHRMMdknX1hKEqsA4Zysd+iXqvJbUqR3wDY/ZCBqaqe4WPrPNXSx0nLeM0MxIkjkwO2Hn9+IKdzB92ZTmTEmrsU+Uo4cUGFL28mNcdDPnskGUk9pqqILy5LAvTehaihfy87szX5n2LyriLG2osZV+oV9ZHJPsVyMonIqKvYfeetr3c4H3XmikgJc2ZlI+O/nW7ap4phzpExocMkEqc8lYKgM3TV2apVtDd0IKaE=',
    'searchtext':'',
    'txtcano':CANumber,
    'txtconsname':Account holder name,
    'txtphone':'',
    'txtmobile':'',
    'txtsuppadd':address,
    'txtremarks': '',
    'TxtImgVer':captcha,
    'btnsubmit': 'Submit'
    }

    try:
        response = requests.post('https://www.tatapower-ddl.com/customer/complaint/regcomp_detailsrevamp.aspx', data = getDetailsPayload)
        response.raise_for_status() #Raises an exception in case of a failed response code
        logging.debug('Response fine')
        #print(response.text)
        soup = BeautifulSoup(response.text, features='html.parser')
        logging.debug('Soup Made')

        address = soup.find(id = 'txtsuppadd').string.strip()
        name = soup.find(id = 'txtconsname')['value'] #https://stackoverflow.com/questions/2612548/extracting-an-attribute-value-with-beautifulsoup
        #phone = soup.find(id = 'txtphone')
        #mobile = soup.find(id = 'txtmobile')
        logging.debug('Scraped fields successfully')

        #Getting remarks from the sheet
        logging.debug('Getting remarks from the sheet')
        remarks = remarksSheet['A']
        
        remarksList = []
        
        for remark in remarks:
            if remark.value != None:
                remarksList.append(str(remark.value))

        remarkCount = len(remarksList)
        
        if remarkCount > 1:
            remark = remarksList[random.randrange(0,remarkCount)]
        elif remarkCount == 1:
            remark = remarksList[0]
        else:
            remark = ''
            
        logging.debug('Remarks fetched')
            
        
        
        #Adding these details to payload
        complaintRegisterPayload['txtsuppadd'] = address
        complaintRegisterPayload['txtconsname'] = name
        complaintRegisterPayload['txtremarks'] = remark
        logging.debug('Updated payload')

        #Updating details in the worksheet
        CASheet.cell(row = index, column = 2).value = name
        CASheet.cell(row = index, column = 5).value = address
        workbook.save(file)
        logging.debug('Saved fields to workbook')

        return(complaintRegisterPayload)
    
    except requests.exceptions.RequestException as e:
        print('Response Error : ', e)


def lodgeComplaint(payload):
    logging.debug('started lodgeComplaint()')

    try:
        '''
        response = requests.post('https://www.tatapower-ddl.com/customer/complaint/regcomp_detailsrevamp.aspx', data = payload)
        response.raise_for_status() #Raises an exception in case of a failed response code
        #TODO : Handle the case if complaint already registered
        logging.debug('Response fine')

        soup = BeautifulSoup(response.text,'html.parser')
        #extract complaint number from response
        '''
        complaintNumber = 'some complaint number'
        
        previousComplaints = complaintSheet['A']
        count = 0

        for complaint in previousComplaints:
            if complaint.value != None:
                count += 1
                
        count += 1

        logging.debug('Adding complaint number in sheet')
        timestamp = datetime.now();timestamp.strftime("%d/%m/%Y %H:%M:%S")
        complaintSheet.cell(row = count, column = 1).value = timestamp
        complaintSheet.cell(row = count, column = 2).value = payload['txtcano']
        complaintSheet.cell(row = count, column = 3).value = 'some complaint number'
        workbook.save(file)
        logging.debug('Finished lodgeComplaint()')
        
    except requests.exceptions.RequestException as e:
        print('Response Error : ', e)

main()


    
