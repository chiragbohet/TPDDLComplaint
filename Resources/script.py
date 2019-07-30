import openpyxl,requests,logging
from datetime import datetime
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

"""
Steps :
1) Get a CA number from sheet
2) For CA number get all the related fields.
3) File a complaint with that CA number
4) Save the complaint number with timestamp to the sheet
5) Repeat steps 1 to 4 for all CA numbers
"""

'''
1) Add a option to randomly pick a complaint description from 3rd sheet. [call rand on a list]
2) Error Handling - if already planned outage etc.
3) 
'''

def lodgeComplaint(CANumber):
    '''
    Lodges a complaint with the CANumber given as input and returns the complaint number
    
    '''
def getDetailsFromCA(CANumber):
    '''
    Gets all the necessary details from the CA Number to lodge a complaint and returns a dictionary of payload
    '''
    logging.debug('Inside - getDetailsFromCA()')
    payload = {
    '__VIEWSTATE':'/wEPDwUILTg3MjIxMDYPZBYCAgEPZBYaZg8PFgIeBFRleHQFgwFPcHRpb24gU2VsZWN0ZWQgOiA8dT5ObyBQb3dlcjwvdT4gKE8mTSAgQ29tcGxhaW50KS4gPGJyPjxhIGhyZWY9J2NvbXBsYWludC1yZWdpc3RyYXRpb24uYXNweCc+Q2xpY2sgaGVyZTwvYT4gdG8gc2VsZWN0IG90aGVyIG9wdGlvbmRkAgIPDxYCHgdWaXNpYmxlZ2RkAgMPDxYEHglCYWNrQ29sb3IKIR4EXyFTQgIIZGQCBQ8PFgIfAWdkZAIGDw8WBh8CCiEeCFJlYWRPbmx5Zx8DAghkZAIHDw9kFgIeCm9ua2V5cHJlc3MFJmphdmFzY3JpcHQ6cmV0dXJuIGFsbG93bnVtYmVycyhldmVudCk7ZAIIDw9kFgIfBQUmamF2YXNjcmlwdDpyZXR1cm4gYWxsb3dudW1iZXJzKGV2ZW50KTtkAgkPDxYCHwFnZGQCCg8PFgYfAgohHwRnHwMCCGRkAg0PEGRkFgBkAg8PDxYCHwFnZGQCEA8PFgQfAgohHwMCCGRkAhEPDxYEHgdUb29sVGlwBRdJbWFnZSBWZXJpZmljYXRpb24gQ29kZR4ISW1hZ2VVcmwFCn4vSW1nLmFzcHhkZGR8HzDRmY0tnIDT84vqDDTUVfzuhoOXvx04gBQKfM+N5Q==',
    '__VIEWSTATEGENERATOR':'FA96F2E7',
    '__EVENTVALIDATION':'/wEdAApYaIhDCcMW/ET5/xEE7sC2EcOgvHRMMdknX1hKEqsA4Zysd+iXqvJbUqR3wDY/ZCBqaqe4WPrPNXSx0nLeM0MxIkjkwO2Hn9+IKdzB92ZTmTEmrsU+Uo4cUGFL28mNcdDPnskGUk9pqqILy5LAvTehaihfy87szX5n2LyriLG2osZV+oV9ZHJPsVyMonIqKvYfeetr3c4H3XmikgJc2ZlI/ecVh9iV+ytFuVNnQebcaFAdHDEasYEiOT9ymVIZlK8=',
    'searchtext':'',
    'txtcano': CANumber,
    'btngetconsdetails':'Get Details',
    'txtconsname':'',
    'txtphone':'',
    'txtmobile':'',
    'txtsuppadd':'',
    'txtremarks':'',
    'TxtImgVer':'Enter code as it appears'
    }
    logging.debug('Constructed Payload')
    
    try:
        response = requests.post('https://www.tatapower-ddl.com/customer/complaint/regcomp_detailsrevamp.aspx',data=payload)
        response.raise_for_status() #Raises an exception in case of a failed response code
        logging.debug('Response fine')  
        soup = BeautifulSoup(response.text,'html.parser')
        logging.debug('Soup made')
        #print(soup.prettify())
        
        #txtremarks this needs to be randomly picked from a list
        print(soup.find(id="txtconsname"))
        print(soup.find(id="txtsuppadd").string)#this works
        #captcha : need to do something about this
        logging.debug('Printed')
    except requests.exceptions.RequestException as e:
        print('Response Error : ', e)
        
getDetailsFromCA('60014463164')

'''
    
file = 'CANumbers.xlsx' #The file name

workbook = openpyxl.load_workbook(filename = file)
CASheet = workbook['CANumbers']
compSheet = workbook['ComplaintNumbers']


compIndex = len(compSheet['A']) + 2 #To leave a space 

for CANumber in CASheet['A']:
    print(CANumber.value) # Process here

    timestamp = datetime.now();timestamp.strftime("%d/%m/%Y %H:%M:%S")
    compSheet.cell(row = compIndex, column = 1, value = timestamp)
    compSheet.cell(row = compIndex, column = 2, value = CANumber.value)
    compIndex += 1
    
    
workbook.save(file)

https://www.tatapower-ddl.com/customer/complaint/complaint-registration.aspx
'''
